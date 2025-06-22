#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
教师信息生成脚本
自动生成教师基本信息并创建 Excel 文件
"""

import random
import os
from faker import Faker
from openpyxl import Workbook
from datetime import datetime

# 初始化 Faker
fake = Faker('zh_CN')

# 学院信息
COLLEGES = [
    "信息与电气工程学院",
    "计算机科学与技术学院",
    "理学院",
    "能源与矿业工程学院",
    "地球科学与测绘工程学院",
    "环境与测绘学院",
    "机电工程学院",
    "材料科学与工程学院",
    "经济管理学院",
    "建筑与设计学院",
    "外国语学院",
    "体育学院",
    "马克思主义学院",
    "数据科学学院"
]

def generate_teacher_no():
    """生成教师工号: T + 7位数字"""
    return f"T{random.randint(1000000, 9999999)}"

def generate_teachers(count):
    """生成指定数量的教师数据"""
    teachers_data = []
    for _ in range(count):
        gender = random.choice([0, 1])  # 0: 女性, 1: 男性
        real_name = fake.name()
        username = fake.user_name()[:10]  # 限制用户名长度为10
        
        teacher = {
            'username': username,
            'realName': real_name,
            'gender': gender,
            'phone': fake.phone_number(),
            'teacherNo': generate_teacher_no(),
            'school': "中国矿业大学",
            'college': random.choice(COLLEGES),
            'password': '123456'  # 默认密码
        }
        teachers_data.append(teacher)
    return teachers_data

def save_to_excel(teachers_data, file_path):
    """将教师数据保存到Excel文件"""
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "教师信息"
    
    # 定义列顺序
    columns = [
        'username', 'realName', 'gender', 'phone', 
        'teacherNo', 'school', 'college', 'password'
    ]
    
    # 添加表头
    for col_num, column_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_name)
    
    # 添加数据
    for row_num, teacher in enumerate(teachers_data, 2):
        for col_num, column_name in enumerate(columns, 1):
            value = teacher.get(column_name)
            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
    
    # 保存文件
    wb.save(file_path)
    print(f"成功生成文件: {file_path}")

def main():
    """主函数"""
    print("欢迎使用教师信息生成脚本！")
    
    try:
        count = int(input("请输入要生成的教师数量: "))
        if count <= 0:
            raise ValueError
    except ValueError:
        print("错误：请输入一个正整数作为教师数量。")
        return
    
    # 可选：让用户指定学院
    use_specific_college = input("是否指定学院？(y/n): ").lower() == 'y'
    if use_specific_college:
        print("可选学院列表:")
        for i, college in enumerate(COLLEGES):
            print(f"{i+1}. {college}")
        
        try:
            college_index = int(input("请输入学院序号: ")) - 1
            if college_index < 0 or college_index >= len(COLLEGES):
                print(f"错误：无效的学院序号，将使用随机学院。")
                college = None
            else:
                college = COLLEGES[college_index]
        except ValueError:
            print("错误：无效的输入，将使用随机学院。")
            college = None
    else:
        college = None
    
    # 生成教师数据
    teachers = generate_teachers(count)
    
    # 如果指定了学院，则设置所有教师的学院
    if college:
        for teacher in teachers:
            teacher['college'] = college
    
    # 确定输出目录和文件名
    output_dir = '../data'
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f"teachers_{timestamp}.xlsx"
    if college:
        file_name = f"{college}_{file_name}"
    
    file_path = os.path.join(output_dir, file_name)
    
    # 保存到Excel
    save_to_excel(teachers, file_path)
    
    print("\n教师信息生成完毕！")

if __name__ == "__main__":
    main()
