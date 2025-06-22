#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
学生信息生成脚本（改进版）
自动生成学生基本信息并创建 Excel 文件
- 确保学号和用户名不重复
- 用户名采用统一格式: stu + 学号
- 生成的 Excel 文件可直接用于管理员批量注册学生接口 (/api/eduagentx/admin/register/student/excel)
"""

import random
import os
from faker import Faker
from openpyxl import Workbook
from datetime import datetime

# 初始化 Faker
fake = Faker('zh_CN')

# 专业信息映射（专业代码：专业名称）
MAJOR_INFO = {
    "01": "采矿工程", "02": "工业工程", "03": "交通运输", "04": "智能采矿工程",
    "05": "新能源科学与工程", "06": "安全工程", "07": "消防工程", "08": "职业卫生工程",
    "09": "应急技术与管理", "10": "土木工程", "11": "工程力学", "12": "建筑环境与能源应用工程",
    "13": "工程管理", "14": "机械工程", "15": "智能制造工程", "16": "机器人工程",
    "17": "电子信息工程", "18": "自动化", "19": "集成电路设计与集成系统", "20": "人工智能",
    "21": "地质工程", "22": "资源勘查工程", "23": "地球物理学", "24": "水文与水资源工程",
    "25": "地球信息科学与技术", "26": "矿物加工工程", "27": "化学工程与工艺", "28": "应用化学",
    "29": "过程装备与控制工程", "30": "生物工程", "31": "能源化学工程", "32": "测绘工程",
    "33": "环境工程", "34": "地理信息科学", "35": "环境科学", "36": "遥感科学与技术",
    "37": "电气工程及其自动化", "38": "能源与动力工程", "39": "储能科学与工程", "40": "材料科学与工程",
    "41": "新能源材料与器件", "42": "应用物理学", "43": "光电信息科学与工程", "44": "数学与应用数学",
    "45": "统计学", "46": "计算机科学与技术", "47": "电子信息科学与技术", "48": "信息安全",
    "49": "数据科学与大数据技术", "50": "软件工程", "51": "会计学", "52": "金融学",
    "53": "国际经济与贸易", "54": "工商管理", "55": "市场营销", "56": "人力资源管理",
    "57": "电子商务", "58": "大数据管理与应用", "59": "行政管理", "60": "土地资源管理",
    "61": "应急管理", "62": "英语", "63": "德语", "64": "建筑学",
    "65": "环境设计", "66": "工业设计", "67": "城乡规划", "68": "汉语言文学",
    "69": "广播电视学", "70": "法学", "71": "音乐学", "72": "网络与新媒体",
    "73": "社会体育指导与管理", "74": "运动训练"
}

# 学院信息
COLLEGES = [
    "信息与电气工程学院",
    "计算机科学与技术学院",
    "理学院",
    "能源与矿业工程学院",
    "地球科学与测绘工程学院",
    "环境与测绘学院",
    "机电工程学院",
    "材料科学与工程学院",
    "经济管理学院",
    "建筑与设计学院",
    "外国语学院",
    "体育学院",
    "马克思主义学院",
    "数据科学学院"
]

def generate_student_no(major_code, grade, counter):
    """
    生成学号: 专业代码(2) + 年级(2) + 序号(4)
    确保学号是唯一的，基于专业、年级和计数器
    """
    return f"{major_code}{grade}{counter:04d}"

def generate_class_code(major_code, grade, class_num):
    """生成班级号: 专业代码(2) + 年级(2) + 班级序号(2)"""
    return f"{major_code}{grade}{class_num:02d}"

def generate_students(major_code, grade, class_count, students_per_class=30):
    """
    为指定的专业和年级生成学生数据
    每个班级 students_per_class 名学生
    """
    all_students = []
    student_counter = 1  # 用于确保学号唯一
    
    for class_num in range(1, class_count + 1):
        class_code = generate_class_code(major_code, grade, class_num)
        college = get_college_for_major(major_code)
        
        for _ in range(students_per_class):
            gender = random.choice([0, 1])  # 0: 女性, 1: 男性
            real_name = fake.name()
            student_no = generate_student_no(major_code, grade, student_counter)
            username = f"stu{student_no}"  # 统一格式：stu + 学号
            
            student = {
                'studentNo': student_no,
                'realName': real_name,
                'gender': gender,
                'username': username,
                'phone': fake.phone_number(),
                'email': f"{username}@example.com",
                'majorCode': major_code,
                'classCode': class_code,
                'college': college,
                'school': "中国矿业大学",
                'password': '123456'  # 默认密码
            }
            all_students.append(student)
            student_counter += 1
    
    return all_students

def get_college_for_major(major_code):
    """根据专业代码返回相应的学院"""
    # 这里使用简单的映射规则，可以根据实际情况调整
    major_code_int = int(major_code)
    if 1 <= major_code_int <= 16:
        return "机电工程学院"
    elif 17 <= major_code_int <= 20:
        return "信息与电气工程学院"
    elif 21 <= major_code_int <= 36:
        return "地球科学与测绘工程学院"
    elif 37 <= major_code_int <= 43:
        return "材料科学与工程学院"
    elif 44 <= major_code_int <= 50:
        return "计算机科学与技术学院"
    elif 51 <= major_code_int <= 61:
        return "经济管理学院"
    elif 62 <= major_code_int <= 63:
        return "外国语学院"
    elif 64 <= major_code_int <= 68:
        return "建筑与设计学院"
    elif 69 <= major_code_int <= 72:
        return "网络与新媒体学院"
    elif 73 <= major_code_int <= 74:
        return "体育学院"
    else:
        return random.choice(COLLEGES)

def save_to_excel(students_data, file_path):
    """将学生数据保存到Excel文件"""
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息"
      # 定义列顺序 - 匹配 AdminController.batchRegisterStudentsFromExcel 期望的格式
    columns = [
        'studentNo', 'realName', 'gender', 'username', 
        'phone', 'email', 'majorCode', 'classCode', 
        'college', 'school', 'password'
    ]
    
    # 添加表头
    for col_num, column_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_name)
    
    # 添加数据
    for row_num, student in enumerate(students_data, 2):
        for col_num, column_name in enumerate(columns, 1):
            value = student.get(column_name)
            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
    
    # 保存文件
    wb.save(file_path)
    print(f"成功生成文件: {file_path}")

def main():
    """主函数"""
    print("欢迎使用学生信息生成脚本（改进版）！")
    
    # 获取用户输入
    major_code = input(f"请输入专业号 (01-74): ")
    if major_code not in MAJOR_INFO:
        print(f"错误：无效的专业号。有效的专业号包括: {', '.join(MAJOR_INFO.keys())}")
        return
        
    grade = input("请输入年级 (例如: 22, 23, 24): ")
    if not (grade.isdigit() and 22 <= int(grade) <= 25):  # 假设年级范围在22-25之间
        print("错误：无效的年级。请输入22-25之间的数字。")
        return

    try:
        class_count = int(input("请输入要生成的班级数量: "))
        if class_count <= 0:
            raise ValueError
    except ValueError:
        print("错误：请输入一个正整数作为班级数量。")
        return
    
    try:
        students_per_class = int(input("请输入每个班级的学生数量 (默认30): ") or "30")
        if students_per_class <= 0:
            raise ValueError
    except ValueError:
        print("错误：请输入一个正整数作为学生数量。将使用默认值30。")
        students_per_class = 30
    
    # 生成学生数据
    major_name = MAJOR_INFO[major_code]
    students = generate_students(major_code, grade, class_count, students_per_class)
    
    # 确定输出目录和文件名
    output_dir = '../data'
    os.makedirs(output_dir, exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f"students_{major_name}_{grade}级_{timestamp}.xlsx"
    file_path = os.path.join(output_dir, file_name)
    
    # 保存到Excel
    save_to_excel(students, file_path)
    
    print(f"\n成功生成 {len(students)} 名学生信息！")
    print(f"专业: {major_name}, 年级: {grade}级, 班级数量: {class_count}")
    print(f"文件保存位置: {file_path}")

if __name__ == "__main__":
    main()
