#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据专业、年级和班级数量生成学生信息脚本
为每个班级生成30名学生，保存到单独的Excel文件中
"""

import random
import os
from faker import Faker
from openpyxl import Workbook
from datetime import datetime

# 初始化 Faker
fake = Faker('zh_CN')

# 专业信息映射（专业代码：专业名称）
MAJOR_INFO = {
    "01": "采矿工程", "02": "工业工程", "03": "交通运输", "04": "智能采矿工程",
    "05": "新能源科学与工程", "06": "安全工程", "07": "消防工程", "08": "职业卫生工程",
    "09": "应急技术与管理", "10": "土木工程", "11": "工程力学", "12": "建筑环境与能源应用工程",
    "13": "工程管理", "14": "机械工程", "15": "智能制造工程", "16": "机器人工程",
    "17": "电子信息工程", "18": "自动化", "19": "集成电路设计与集成系统", "20": "人工智能",
    "21": "地质工程", "22": "资源勘查工程", "23": "地球物理学", "24": "水文与水资源工程",
    "25": "地球信息科学与技术", "26": "矿物加工工程", "27": "化学工程与工艺", "28": "应用化学",
    "29": "过程装备与控制工程", "30": "生物工程", "31": "能源化学工程", "32": "测绘工程",
    "33": "环境工程", "34": "地理信息科学", "35": "环境科学", "36": "遥感科学与技术",
    "37": "电气工程及其自动化", "38": "能源与动力工程", "39": "储能科学与工程", "40": "材料科学与工程",
    "41": "新能源材料与器件", "42": "应用物理学", "43": "光电信息科学与工程", "44": "数学与应用数学",
    "45": "统计学", "46": "计算机科学与技术", "47": "电子信息科学与技术", "48": "信息安全",
    "49": "数据科学与大数据技术", "50": "软件工程", "51": "会计学", "52": "金融学",
    "53": "国际经济与贸易", "54": "工商管理", "55": "市场营销", "56": "人力资源管理",
    "57": "电子商务", "58": "大数据管理与应用", "59": "行政管理", "60": "土地资源管理",
    "61": "应急管理", "62": "英语", "63": "德语", "64": "建筑学",
    "65": "环境设计", "66": "工业设计", "67": "城乡规划", "68": "汉语言文学",
    "69": "广播电视学", "70": "法学", "71": "音乐学", "72": "网络与新媒体",
    "73": "社会体育指导与管理", "74": "运动训练"
}

def generate_student_no(major_code, grade):
    """生成8位学号: 专业代码(2) + 年级(2) + 随机4位数"""
    return f"{major_code}{grade}{random.randint(1000, 9999)}"

def generate_class_code(major_code, grade, class_num):
    """生成8位班级号: 专业代码(2) + 年级(2) + 随机4位数"""
    return f"{major_code}{grade}{random.randint(1000, 9999)}"

def generate_students_for_class(major_code, grade, class_code):
    """为一个班级生成30名学生数据"""
    students_data = []
    for _ in range(30):
        real_name = fake.name()
        username = fake.user_name()[:10]  # 限制用户名长度为10
        student_no = generate_student_no(major_code, grade)
        
        student = {
            'studentNo': student_no,
            'realName': real_name,
            'gender': random.choice([0, 1]),
            'username': username,
            'phone': fake.phone_number(),
            'email': fake.email(),
            'majorCode': major_code,
            'classCode': class_code,
            'college': "信息与电气工程学院",  # 示例学院
            'school': "中国矿业大学",
            'createTime': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'updateTime': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'deleteTime': None,
            'tag': True,
            'password': '123456'  # 默认密码
        }
        students_data.append(student)
    return students_data

def save_to_excel(students_data, columns, file_path):
    """将学生数据保存到Excel文件"""
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    
    # 添加表头
    for col_num, column_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_name)
    
    # 添加数据
    for row_num, student in enumerate(students_data, 2):
        for col_num, column_name in enumerate(columns, 1):
            value = student.get(column_name)
            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
    
    # 保存文件
    wb.save(file_path)

def main():
    """主函数"""
    print("欢迎使用学生信息生成脚本！")
    
    # 获取用户输入
    major_code = input(f"请输入专业号 (01-74): ")
    if major_code not in MAJOR_INFO:
        print(f"错误：无效的专业号。有效的专业号包括: {', '.join(MAJOR_INFO.keys())}")
        return
        
    grade = input("请输入年级 (例如: 22, 23, 24): ")
    if not (grade.isdigit() and 22 <= int(grade) <= 25):  # 假设年级范围在22-25之间
        print("错误：无效的年级。请输入22-25之间的数字。")
        return

    try:
        num_classes = int(input("请输入要生成的班级数量: "))
        if num_classes <= 0:
            raise ValueError
    except ValueError:
        print("错误：请输入一个正整数作为班级数量。")
        return

    major_name = MAJOR_INFO[major_code]
    output_dir = '../data'
    os.makedirs(output_dir, exist_ok=True)

    for i in range(1, num_classes + 1):
        class_code = generate_class_code(major_code, grade, i)
        print(f"正在为班级 {class_code} ({major_name}) 生成学生数据...")
        
        students = generate_students_for_class(major_code, grade, class_code)
        
        # 定义列顺序
        columns_order = [
            'username', 'realName', 'gender', 'phone', 'email', 
            'majorCode', 'classCode', 'college', 'school', 
            'createTime', 'updateTime', 'deleteTime', 'tag', 'password'
        ]

        file_name = f"{class_code}-{major_name}.xlsx"
        file_path = os.path.join(output_dir, file_name)
        
        save_to_excel(students, columns_order, file_path)
        print(f"成功生成文件: {file_path}")

    print("\n所有文件生成完毕！")

if __name__ == "__main__":
    main()
